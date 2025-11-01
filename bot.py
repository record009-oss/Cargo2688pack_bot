import logging
import pandas as pd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler
from io import BytesIO
import openpyxl
import openpyxl.styles
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import tempfile
import os

# ВСТАВЬ СВОЙ ТОКЕН ЗДЕСЬ
BOT_TOKEN = "8346614759:AAHbqo5tm34zlVyNmy4_0k_suxe3dgG93ks"

# Состояния диалога
CLIENT_CODE, TRACK_NUMBER, PHOTO, NAME, COLOR_SIZE, QUANTITY, LINK, ADD_MORE = range(8)

# Хранилище данных пользователей
user_data = {}

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# Установка меню команд при запуске бота
async def post_init(application: Application):
    await application.bot.set_my_commands([
        ("start", "Начать работу с ботом")
    ])

# Команда /start - основная точка входа
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_name = update.effective_user.first_name
    
    keyboard = [
        [InlineKeyboardButton("🚀 Начать работу", callback_data="start_work")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        f"Привет, {user_name}👋 !\n\n"
        "Добро пожаловать в бот Cargo_2688 для создания упаковочного листа\n\n"
        "📦 Этот бот поможет:\n"
        "• 📸 Создавать упаковочные листы с фото товаров и ссылкой на товар\n"
        "• 📋 Экспортировать данные в Excel\n\n"
        "Для продолжения нажмите - начать работу",
        reply_markup=reply_markup
    )

# Обработка нажатия инлайн кнопки "Начать работу"
async def start_work(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    user_data[user_id] = {
        'client_code': None,  # Код клиента (только один раз)
        'positions': []       # Список для хранения нескольких позиций
    }
    
    await query.edit_message_text(
        "🚀 Начинаем создание упаковочного листа!\n\n"
        "1️⃣ Введите код клиента:"
    )
    
    return CLIENT_CODE

# Обработка кнопки "Создать новый заказ" после отправки таблицы
async def new_order_after_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    user_data[user_id] = {
        'client_code': None,
        'positions': []
    }
    
    # Отправляем новое сообщение вместо редактирования старого
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="🚀 Начинаем создание упаковочного листа!\n\n"
             "1️⃣ Введите код клиента:"
    )
    
    return CLIENT_CODE

# Обработка кода клиента
async def handle_client_code(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_data[user_id]['client_code'] = update.message.text
    
    await update.message.reply_text(
        "✅ Код клиента сохранен!\n\n"
        "2️⃣ Напишите Трэк номер для первого товара:"
    )
    
    return TRACK_NUMBER

# Обработка трек номера
async def handle_track_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Сохраняем текущий трек номер во временное хранилище
    context.user_data['current_track_number'] = update.message.text
    
    await update.message.reply_text(
        "✅ Трэк номер сохранен!\n\n"
        "3️⃣ Отправьте фото товара (jpg):"
    )
    
    return PHOTO

# Обработка фото
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Сохраняем информацию о фото
    photo = update.message.photo[-1]
    current_photo = {
        'file_id': photo.file_id,
        'file_unique_id': photo.file_unique_id
    }
    
    # Сохраняем текущее фото во временное хранилище
    context.user_data['current_photo'] = current_photo
    
    await update.message.reply_text(
        "✅ Фото сохранено!\n\n"
        "4️⃣ Введите название товара:"
    )
    
    return NAME

# Обработка названия
async def handle_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    context.user_data['current_name'] = update.message.text
    
    await update.message.reply_text(
        "✅ Название сохранено!\n\n"
        "5️⃣ Введите Цвет и размер:\n"
        "Пример: \"красный - L, желтый - 40\""
    )
    
    return COLOR_SIZE

# Обработка цвета и размеров
async def handle_color_size(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    context.user_data['current_color_size'] = update.message.text
    
    await update.message.reply_text(
        "✅ Цвет и размеры сохранены!\n\n"
        "6️⃣ Введите количество товара:"
    )
    
    return QUANTITY

# Обработка количества и переход к ссылке
async def handle_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    context.user_data['current_quantity'] = update.message.text
    
    # Создаем клавиатуру с кнопкой "Пропустить"
    keyboard = [
        [InlineKeyboardButton("⏭️ Пропустить", callback_data="skip_link")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "✅ Количество сохранено!\n\n"
        "7️⃣ Пришлите ссылку на товар\n"
        "*если нет ссылки нажмите пропустить",
        reply_markup=reply_markup
    )
    
    return LINK

# Обработка ссылки на товар
async def handle_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    context.user_data['current_link'] = update.message.text
    return await save_position_and_continue(update, context, user_id)

# Обработка кнопки "Пропустить"
async def skip_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    context.user_data['current_link'] = "Нет ссылки"
    return await save_position_and_continue(update, context, user_id, query)

# Общая функция для сохранения позиции
async def save_position_and_continue(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, query=None):
    current_position = {
        'track_number': context.user_data.get('current_track_number'),
        'photo': context.user_data.get('current_photo'),
        'name': context.user_data.get('current_name'),
        'color_size': context.user_data.get('current_color_size'),
        'quantity': context.user_data.get('current_quantity'),
        'link': context.user_data.get('current_link', 'Нет ссылки')
    }
    
    user_data[user_id]['positions'].append(current_position)
    
    # Очищаем временные данные
    context.user_data.pop('current_track_number', None)
    context.user_data.pop('current_photo', None)
    context.user_data.pop('current_name', None)
    context.user_data.pop('current_color_size', None)
    context.user_data.pop('current_quantity', None)
    context.user_data.pop('current_link', None)
    
    keyboard = [
        [InlineKeyboardButton("➕ Добавить еще товары", callback_data="add_more")],
        [InlineKeyboardButton("✅ Завершить и отправить", callback_data="finish")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    position_count = len(user_data[user_id]['positions'])
    
    if query:
        await query.edit_message_text(
            f"✅ Товар добавлен!\n\n"
            f"📦 Всего товаров: {position_count}\n\n"
            "Выберите действие:",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            f"✅ Товар добавлен!\n\n"
            f"📦 Всего товаров: {position_count}\n\n"
            "Выберите действие:",
            reply_markup=reply_markup
        )
    
    return ADD_MORE

# Обработка кнопки "Добавить ещё товары"
async def add_more_position(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "🔄 Добавляем новый товар!\n\n"
        "2️⃣ Напишите Трэк номер для этого товара:"
    )
    
    return TRACK_NUMBER

# Обработка кнопки "Завершить и отправить"
async def finish_and_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    position_count = len(user_data[user_id]['positions'])
    
    await query.edit_message_text(
        f"📊 Формируем упаковочный лист...\n\n"
        f"📦 Товаров: {position_count}\n"
        f"⏳ Пожалуйста, подождите..."
    )
    
    await create_and_send_table(update, context, user_id)
    return ConversationHandler.END

# Создание и отправка таблицы
async def create_and_send_table(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int):
    data = user_data[user_id]
    positions = data['positions']
    
    output = BytesIO()
    temp_file_paths = []
    
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Упаковочный лист'
        
        headers = ['Код клиента', 'Трэк номер', 'Фото товара', 'Наименование', 'Цвет и размер', 'Количество', 'Ссылка', '№']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            cell.fill = openpyxl.styles.PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        for idx, position in enumerate(positions, 1):
            row = idx + 1
            
            worksheet.cell(row=row, column=1, value=data['client_code'])
            worksheet.cell(row=row, column=2, value=position['track_number'])
            worksheet.cell(row=row, column=4, value=position['name'])
            worksheet.cell(row=row, column=5, value=position['color_size'])
            worksheet.cell(row=row, column=6, value=position['quantity'])
            worksheet.cell(row=row, column=7, value=position['link'])
            worksheet.cell(row=row, column=8, value=idx)
            
            if position['photo']:
                try:
                    file = await context.bot.get_file(position['photo']['file_id'])
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as tmp_file:
                        temp_file_path = tmp_file.name
                    temp_file_paths.append(temp_file_path)
                    
                    await file.download_to_drive(temp_file_path)
                    
                    if os.path.exists(temp_file_path) and os.path.getsize(temp_file_path) > 0:
                        img = Image(temp_file_path)
                        img.width = 80
                        img.height = 80
                        worksheet.add_image(img, f'C{row}')
                        worksheet.row_dimensions[row].height = 65
                        
                except Exception as e:
                    print(f"❌ Ошибка добавления фото для позиции {idx}: {e}")
                    worksheet.cell(row=row, column=3, value="Ошибка фото")
            else:
                worksheet.cell(row=row, column=3, value="Нет фото")
        
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 30
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 40
        worksheet.column_dimensions['H'].width = 8
        
        for row in range(2, len(positions) + 2):
            for col in [1, 2, 4, 5, 6, 7, 8]:
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = openpyxl.styles.Alignment(vertical='center')
        
        worksheet.auto_filter.ref = f"A1:H{len(positions) + 1}"
        workbook.save(output)
        
    except Exception as e:
        print(f"❌ Критическая ошибка создания Excel: {e}")
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        simple_data = []
        for idx, position in enumerate(positions, 1):
            simple_data.append({
                '№': idx,
                'Код клиента': data['client_code'],
                'Трэк номер': position['track_number'],
                'Наименование': position['name'],
                'Цвет и размер': position['color_size'],
                'Количество': position['quantity'],
                'Ссылка': position['link']
            })
        
        df = pd.DataFrame(simple_data)
        df.to_excel(output, index=False, engine='openpyxl')
    
    finally:
        for temp_file_path in temp_file_paths:
            if os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except Exception as e:
                    print(f"⚠️ Не удалось удалить временный файл: {e}")
    
    output.seek(0)
    
    # Создаем кнопку для нового заказа
    keyboard = [
        [InlineKeyboardButton("🔄 Создать новый заказ", callback_data="new_order_after_finish")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    position_count = len(positions)
    await context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=output,
        filename=f"упаковочный_лист.xlsx",
        caption=(
            f"✅ Таблица успешно создана!\n\n"
            f"📦 Количество позиций: {position_count}\n\n"
            "Чтобы начать новый заказ нажмите:"
        ),
        reply_markup=reply_markup
    )
    
    user_data[user_id] = {}

# Отмена диалога
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in user_data:
        del user_data[user_id]
    
    await update.message.reply_text(
        "Диалог отменен. Чтобы начать заново, нажмите /start"
    )
    return ConversationHandler.END

# Обработка любых текстовых сообщений
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_name = update.effective_user.first_name
    await update.message.reply_text(
        f"Привет, {user_name}👋 !\n\n"
        "Для создания упаковочного листа используйте команду /start"
    )

def main():
    application = Application.builder().token(BOT_TOKEN).build()
    
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("cancel", cancel))
    
    conv_handler = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(start_work, pattern='^start_work$'),
            CallbackQueryHandler(new_order_after_finish, pattern='^new_order_after_finish$')
        ],
        states={
            CLIENT_CODE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_client_code)],
            TRACK_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_track_number)],
            PHOTO: [MessageHandler(filters.PHOTO, handle_photo)],
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_name)],
            COLOR_SIZE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_color_size)],
            QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_quantity)],
            LINK: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_link),
                CallbackQueryHandler(skip_link, pattern='^skip_link$')
            ],
            ADD_MORE: [
                CallbackQueryHandler(add_more_position, pattern='^add_more$'),
                CallbackQueryHandler(finish_and_send, pattern='^finish$')
            ]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    application.add_handler(conv_handler)
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    application.post_init = post_init
    
    print("🤖 Бот запущен! Проверь его в Telegram.")
    application.run_polling()

if __name__ == "__main__":
    main()